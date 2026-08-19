"""Export smoke tests: the workbook and the web payload.

These check structure rather than content. The point is that a change to
the plan cannot silently break the two artefacts anyone outside the
engine actually sees.
"""

import json

import pytest
from openpyxl import load_workbook

from dcww_planning.excelpack import export_planning_pack
from dcww_planning.planner import build_plan, run_scenarios
from dcww_planning.risk import build_risk_register
from dcww_planning.webexport import build_payload, write_payload

HORIZON = 56


@pytest.fixture(scope="module")
def artefacts(config, daily, profiles):
    plan = build_plan(config, daily, profiles, horizon_days=HORIZON)
    risks = build_risk_register(plan, config)
    table, _ = run_scenarios(
        config, daily, profiles, horizon_days=HORIZON,
        scenarios={"base": {}, "volume_up_10": {"volume_multiplier": 1.10}},
    )
    return plan, risks, table


def test_workbook_has_every_expected_sheet(artefacts, config, tmp_path):
    plan, risks, table = artefacts
    path = export_planning_pack(plan, config, tmp_path / "pack.xlsx",
                                risks=risks, scenarios=table)
    assert path.exists() and path.stat().st_size > 10_000

    wb = load_workbook(path)
    expected = {
        "Summary", "Assumptions", "Weekly plan", "Supply & recruitment",
        "Daily forecast", "Interval requirement", "Scenarios",
        "Forecast accuracy", "Risk register",
    }
    assert expected <= set(wb.sheetnames)


def test_assumptions_sheet_carries_live_formulas(artefacts, config, tmp_path):
    """The pack must be a model a planner can argue with, not a printout.

    If shrinkage is exported as a hard number, changing sickness means
    re-running Python and re-exporting - which is exactly the round trip
    that stops people using a tool.
    """
    plan, risks, table = artefacts
    path = export_planning_pack(plan, config, tmp_path / "pack.xlsx",
                                risks=risks, scenarios=table)
    ws = load_workbook(path)["Assumptions"]
    formulas = [
        cell.value for row in ws.iter_rows() for cell in row
        if isinstance(cell.value, str) and cell.value.startswith("=")
    ]
    assert formulas, "no live formulas found on the assumptions sheet"
    assert any("(1-B" in f for f in formulas), "shrinkage is not compounded in Excel"
    assert any(f.startswith("=1/(1-B") for f in formulas), "no uplift factor formula"


def test_weekly_sheet_row_count_matches_the_plan(artefacts, config, tmp_path):
    plan, risks, table = artefacts
    path = export_planning_pack(plan, config, tmp_path / "pack.xlsx",
                                risks=risks, scenarios=table)
    ws = load_workbook(path)["Weekly plan"]
    header_row = next(r for r in range(1, 12) if ws.cell(row=r, column=1).value == "Week Start")
    data_rows = sum(1 for r in range(header_row + 1, ws.max_row + 1)
                    if ws.cell(row=r, column=1).value is not None)
    assert data_rows == len(plan.weekly)


def test_web_payload_is_json_serialisable_and_complete(artefacts, config, daily, tmp_path):
    plan, risks, table = artefacts
    payload = build_payload(plan, config, daily, risks=risks, scenarios=table)

    for key in ("meta", "config", "headline", "weekly", "history",
                "forecastWeekly", "intraday", "accuracy", "supply", "scenarios", "risks"):
        assert key in payload, key

    path = write_payload(payload, tmp_path / "payload.json")
    reloaded = json.loads(path.read_text())
    assert reloaded["meta"]["weeks"] == len(plan.weekly)
    assert len(reloaded["weekly"]) == len(plan.weekly)
    assert reloaded["config"]["channels"]
    assert reloaded["config"]["shrinkage"]["total"] > 0


def test_payload_carries_the_inputs_needed_for_client_side_recompute(artefacts, config, daily):
    """The demo recomputes Erlang in the browser, so it needs inputs.

    Shipping only the finished FTE numbers would leave the sliders with
    nothing to act on and reduce the demo to a picture.
    """
    plan, risks, table = artefacts
    payload = build_payload(plan, config, daily, risks=risks, scenarios=table)

    assert payload["intraday"], "no interval detail for client-side sizing"
    for line in payload["intraday"]:
        assert len(line["contacts"]) == 48
        assert line["ahtSeconds"] > 0
        assert line["channel"]

    channels = {c["key"]: c for c in payload["config"]["channels"]}
    voice = channels["voice"]
    assert voice["kind"] == "interactive"
    assert voice["slTarget"] > 0 and voice["slSeconds"] > 0
    assert voice["maxOccupancy"] > 0
    assert payload["config"]["cost"]["hourly"] > 0
    assert payload["config"]["supply"]["pipelineWeeks"] > 0


def test_payload_states_the_data_is_synthetic(artefacts, config, daily):
    """Non-negotiable. The provenance travels with the artefact."""
    plan, risks, table = artefacts
    payload = build_payload(plan, config, daily, risks=risks, scenarios=table)
    note = payload["meta"]["dataNote"].lower()
    assert "synthetic" in note
