"""Excel planning pack export.

"Good knowledge of Excel is essential." Excel is not a lesser output here -
it is the medium the operation actually reviews plans in, and a Python
engine that cannot hand its answer to a planner in a workbook they can
interrogate has solved the easy half of the problem.

Two decisions make this a working model rather than a data dump:

* **The assumptions sheet carries live formulas.** Shrinkage is written as
  real Excel arithmetic referencing the component cells, so a planner can
  change sickness from 4.5% to 6% and watch the total, the uplift factor
  and the requirement move. A pasted number would have to be recalculated
  in Python and re-exported, which is exactly the round trip that stops
  people using a model.
* **Every sheet is formatted for reading, not for storage** - frozen
  panes, number formats, conditional formatting on the gap column - so
  the week that breaks is visible at a glance rather than found by
  scrolling.
"""

from __future__ import annotations

from datetime import date
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .config import PlanConfig

__all__ = ["export_planning_pack"]

BRAND = "1F4E5F"        # deep teal - water utility, not a corporate blue
ACCENT = "2E8B92"
LIGHT = "E8F1F2"
WARN = "FFF2CC"
BAD = "F8CBCB"
GOOD = "D9EAD3"

HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
HEADER_FILL = PatternFill("solid", fgColor=BRAND)
TITLE_FONT = Font(bold=True, size=14, color=BRAND)
SUB_FONT = Font(italic=True, size=9, color="666666")
THIN = Side(style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _write_frame(ws, frame: pd.DataFrame, start_row: int = 1, number_formats: dict | None = None) -> int:
    """Write a DataFrame with a styled header. Returns the last row used."""
    number_formats = number_formats or {}
    for c, name in enumerate(frame.columns, start=1):
        cell = ws.cell(row=start_row, column=c, value=str(name).replace("_", " ").title())
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BOX

    for r, row in enumerate(frame.itertuples(index=False), start=start_row + 1):
        for c, value in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=value)
            cell.border = BOX
            column = frame.columns[c - 1]
            if column in number_formats:
                cell.number_format = number_formats[column]

    for c, name in enumerate(frame.columns, start=1):
        longest = max([len(str(name))] + [len(str(v)) for v in frame.iloc[:, c - 1].head(200)])
        ws.column_dimensions[get_column_letter(c)].width = min(max(longest + 3, 11), 62)

    ws.freeze_panes = ws.cell(row=start_row + 1, column=1)
    return start_row + len(frame)


def _title(ws, text: str, subtitle: str = "") -> int:
    ws["A1"] = text
    ws["A1"].font = TITLE_FONT
    row = 2
    if subtitle:
        ws["A2"] = subtitle
        ws["A2"].font = SUB_FONT
        row = 3
    return row + 1


def export_planning_pack(
    plan,
    config: PlanConfig,
    path: str | Path,
    risks: list | None = None,
    scenarios: pd.DataFrame | None = None,
) -> Path:
    """Write the full planning pack workbook."""
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()

    _sheet_summary(wb.active, plan, config, risks)
    _sheet_assumptions(wb.create_sheet("Assumptions"), config)
    _sheet_weekly(wb.create_sheet("Weekly plan"), plan)
    _sheet_supply(wb.create_sheet("Supply & recruitment"), plan)
    _sheet_forecast(wb.create_sheet("Daily forecast"), plan)
    _sheet_intervals(wb.create_sheet("Interval requirement"), plan)
    if scenarios is not None and not scenarios.empty:
        _sheet_scenarios(wb.create_sheet("Scenarios"), scenarios)
    _sheet_accuracy(wb.create_sheet("Forecast accuracy"), plan)
    if risks:
        _sheet_risks(wb.create_sheet("Risk register"), risks)

    wb.save(path)
    return path


def _sheet_summary(ws, plan, config: PlanConfig, risks) -> None:
    ws.title = "Summary"
    head = plan.headline()
    row = _title(
        ws,
        "Retail Resource Plan - Planning & Performance",
        f"{config.name}  |  horizon {head['horizon_start']:%d %b %Y} to {head['horizon_end']:%d %b %Y}  "
        f"|  all figures from synthetic data",
    )

    metrics = [
        ("Planning weeks", head["weeks"], "0"),
        ("Forecast contacts", head["total_contacts"], "#,##0"),
        ("Mean FTE required", head["mean_required_fte"], "#,##0.0"),
        ("Peak FTE required", head["peak_required_fte"], "#,##0.0"),
        ("Peak week", head["peak_week"], None),
        ("Weeks below required cover", head["weeks_short"], "0"),
        ("Worst weekly gap (FTE)", head["worst_gap_fte"], "#,##0.0"),
        ("Total resourcing cost", head["total_cost"], '"GBP "#,##0'),
        ("Of which overtime + agency premium", head["premium_cost"], '"GBP "#,##0'),
        ("Shrinkage applied (compounded)", config.shrinkage.total, "0.0%"),
        ("Shrinkage uplift factor", config.shrinkage.uplift_factor, "0.000"),
        ("Opening FTE", config.supply.opening_fte, "#,##0.0"),
        ("Annual attrition", config.supply.annual_attrition, "0.0%"),
        ("Recruitment pipeline (weeks)", config.supply.recruitment_lead_weeks
         + config.supply.training_weeks + config.supply.nesting_weeks, "0"),
        ("Heads booked in intake plan", sum(plan.intake.values()), "0"),
    ]
    ws.cell(row=row, column=1, value="Headline").font = Font(bold=True, size=11, color=BRAND)
    row += 1
    for label, value, fmt in metrics:
        ws.cell(row=row, column=1, value=label).border = BOX
        cell = ws.cell(row=row, column=2, value=value)
        cell.border = BOX
        if fmt:
            cell.number_format = fmt
        ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor=LIGHT)
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="How this pack maps to the role").font = Font(bold=True, size=11, color=BRAND)
    row += 1
    mapping = [
        ("Short / medium / long term forecasts and resource plans", "Weekly plan, Daily forecast, Scenarios"),
        ("Bespoke forecasting models, KPIs met cost-effectively", "Forecast accuracy, Scenarios"),
        ("WFM/WFO operation - forecasting and scheduling", "Interval requirement, Supply & recruitment"),
        ("Contact centre metrics and trend investigation", "Forecast accuracy, Interval requirement"),
        ("Pan-Retail, multi-channel plan", "Weekly plan (voice, chat, email, social, back office)"),
        ("Real Time Analysis, adherence, off-line exceptions", "RTA report (separate export)"),
        ("Identify future issues and propose mitigation", "Risk register"),
    ]
    for requirement, where in mapping:
        ws.cell(row=row, column=1, value=requirement).border = BOX
        ws.cell(row=row, column=2, value=where).border = BOX
        row += 1

    ws.column_dimensions["A"].width = 52
    ws.column_dimensions["B"].width = 46

    if risks:
        row += 1
        ws.cell(row=row, column=1, value="Top risks").font = Font(bold=True, size=11, color=BRAND)
        row += 1
        for risk in risks[:5]:
            ws.cell(row=row, column=1, value=f"[{risk.severity.upper()}] {risk.title}").border = BOX
            ws.cell(row=row, column=2, value=risk.mitigation[:150]).border = BOX
            row += 1


def _sheet_assumptions(ws, config: PlanConfig) -> None:
    """Assumptions with live formulas, so the model can be argued with."""
    row = _title(
        ws, "Planning assumptions",
        "Shaded cells are inputs. Everything else is a live formula - change an input and the pack recalculates.",
    )

    ws.cell(row=row, column=1, value="Shrinkage build-up").font = Font(bold=True, color=BRAND)
    row += 1
    ws.cell(row=row, column=1, value="Component").font = HEADER_FONT
    ws.cell(row=row, column=1).fill = HEADER_FILL
    ws.cell(row=row, column=2, value="Rate").font = HEADER_FONT
    ws.cell(row=row, column=2).fill = HEADER_FILL
    ws.cell(row=row, column=3, value="Type").font = HEADER_FONT
    ws.cell(row=row, column=3).fill = HEADER_FILL
    row += 1

    first = row
    regular_rows, irregular_rows = [], []
    kinds = {
        "Annual leave": "Regular", "Training": "Regular", "Coaching / 1-2-1": "Regular",
        "Team meetings": "Regular", "Paid breaks": "Regular",
        "Sickness": "Irregular", "System downtime": "Irregular", "Other offline": "Irregular",
    }
    for name, value in config.shrinkage.components().items():
        ws.cell(row=row, column=1, value=name).border = BOX
        cell = ws.cell(row=row, column=2, value=value)
        cell.number_format = "0.00%"
        cell.fill = PatternFill("solid", fgColor=WARN)
        cell.border = BOX
        kind = kinds.get(name, "Regular")
        ws.cell(row=row, column=3, value=kind).border = BOX
        (regular_rows if kind == "Regular" else irregular_rows).append(row)
        row += 1

    def product_formula(rows: list[int]) -> str:
        return "1-(" + "*".join(f"(1-B{r})" for r in rows) + ")"

    row += 1
    reg_row = row
    ws.cell(row=row, column=1, value="Regular shrinkage (compounded)").font = Font(bold=True)
    ws.cell(row=row, column=2, value=f"={product_formula(regular_rows)}").number_format = "0.00%"
    row += 1
    irr_row = row
    ws.cell(row=row, column=1, value="Irregular shrinkage (compounded)").font = Font(bold=True)
    ws.cell(row=row, column=2, value=f"={product_formula(irregular_rows)}").number_format = "0.00%"
    row += 1
    tot_row = row
    ws.cell(row=row, column=1, value="Total shrinkage").font = Font(bold=True)
    ws.cell(row=row, column=2, value=f"=1-((1-B{reg_row})*(1-B{irr_row}))").number_format = "0.00%"
    ws.cell(row=row, column=2).fill = PatternFill("solid", fgColor=GOOD)
    row += 1
    ws.cell(row=row, column=1, value="Uplift factor (rostered / on phone)").font = Font(bold=True)
    ws.cell(row=row, column=2, value=f"=1/(1-B{tot_row})").number_format = "0.000"
    ws.cell(row=row, column=2).fill = PatternFill("solid", fgColor=GOOD)
    row += 1
    ws.cell(row=row, column=1, value="Naive sum of components (the common error)")
    ws.cell(row=row, column=2, value=f"=SUM(B{first}:B{first + len(kinds) - 1})").number_format = "0.00%"
    row += 1
    ws.cell(row=row, column=1, value="Difference - FTE understated per 100 on phone")
    ws.cell(row=row, column=2, value=f"=100/(1-B{tot_row})-100*(1+B{row - 1})").number_format = "0.0"
    row += 2

    ws.cell(row=row, column=1, value="Channel assumptions").font = Font(bold=True, color=BRAND)
    row += 1
    frame = pd.DataFrame([{
        "channel": c.label, "type": c.kind, "aht_seconds": c.aht_seconds,
        "concurrency": c.concurrency,
        "service_level": f"{c.service_level_target:.0%} in {c.service_level_seconds:.0f}s"
        if c.kind == "interactive" else f"{c.sla_hours:.0f}h SLA",
        "max_occupancy": c.max_occupancy if c.kind == "interactive" else None,
        "productive_utilisation": c.productive_utilisation if c.kind == "deferrable" else None,
        "mean_patience_seconds": c.patience_seconds if c.kind == "interactive" else None,
    } for c in config.channels])
    _write_frame(ws, frame, start_row=row)
    ws.freeze_panes = None

    row += len(frame) + 3
    ws.cell(row=row, column=1, value="Cost and supply").font = Font(bold=True, color=BRAND)
    row += 1
    for label, value, fmt in [
        ("Advisor hourly cost (fully loaded)", config.cost.advisor_hourly_cost, '"GBP "#,##0.00'),
        ("Overtime multiplier", config.cost.overtime_multiplier, "0.00"),
        ("Agency multiplier", config.cost.agency_multiplier, "0.00"),
        ("Recruitment cost per head", config.cost.recruitment_cost_per_head, '"GBP "#,##0'),
        ("Contracted hours per week", config.cost.contracted_hours_per_week, "0.0"),
        ("Opening FTE", config.supply.opening_fte, "#,##0.0"),
        ("Annual attrition", config.supply.annual_attrition, "0.0%"),
        ("Training weeks", config.supply.training_weeks, "0"),
        ("Nesting weeks", config.supply.nesting_weeks, "0"),
        ("Nesting productivity", config.supply.nesting_productivity, "0%"),
        ("Recruitment lead time (weeks)", config.supply.recruitment_lead_weeks, "0"),
        ("Max intake per month", config.supply.max_intake_per_month, "0"),
    ]:
        ws.cell(row=row, column=1, value=label).border = BOX
        cell = ws.cell(row=row, column=2, value=value)
        cell.number_format = fmt
        cell.fill = PatternFill("solid", fgColor=WARN)
        cell.border = BOX
        row += 1

    ws.column_dimensions["A"].width = 46
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 14


def _sheet_weekly(ws, plan) -> None:
    row = _title(ws, "Weekly resource plan", "Gap = supply minus requirement. Red is short, green is surplus.")
    frame = plan.weekly[[
        "week_start", "contacts", "workload_hours", "on_phone_fte", "required_fte",
        "supply_fte", "gap_fte", "coverage_pct", "intake", "leavers", "in_training",
        "overtime_hours", "agency_hours", "idle_hours",
        "base_cost", "overtime_cost", "agency_cost", "recruitment_cost", "total_cost",
    ]].copy()
    last = _write_frame(ws, frame, start_row=row, number_formats={
        "contacts": "#,##0", "workload_hours": "#,##0", "on_phone_fte": "#,##0.0",
        "required_fte": "#,##0.0", "supply_fte": "#,##0.0", "gap_fte": "#,##0.0",
        "coverage_pct": "0.0", "leavers": "0.0", "in_training": "0.0",
        "overtime_hours": "#,##0", "agency_hours": "#,##0", "idle_hours": "#,##0",
        "base_cost": '"GBP "#,##0', "overtime_cost": '"GBP "#,##0',
        "agency_cost": '"GBP "#,##0', "recruitment_cost": '"GBP "#,##0',
        "total_cost": '"GBP "#,##0',
    })
    gap = f"G{row + 1}:G{last}"
    ws.conditional_formatting.add(gap, CellIsRule(
        operator="lessThan", formula=["-0.5"], fill=PatternFill("solid", fgColor=BAD)))
    ws.conditional_formatting.add(gap, CellIsRule(
        operator="greaterThan", formula=["0.5"], fill=PatternFill("solid", fgColor=GOOD)))
    ws.conditional_formatting.add(f"E{row + 1}:E{last}", ColorScaleRule(
        start_type="min", start_color="FFFFFF", end_type="max", end_color=ACCENT))


def _sheet_supply(ws, plan) -> None:
    row = _title(
        ws, "Supply and recruitment plan",
        "Recruits are not capacity until they clear training and nesting - the pipeline is "
        f"{plan.config.supply.recruitment_lead_weeks + plan.config.supply.training_weeks + plan.config.supply.nesting_weeks} weeks.",
    )
    frame = pd.DataFrame([s.as_row() for s in plan.supply])
    _write_frame(ws, frame, start_row=row, number_formats={
        "opening_fte": "#,##0.0", "leavers": "0.00", "in_training": "#,##0.0",
        "nesting": "#,##0.0", "productive_fte": "#,##0.0",
    })


def _sheet_forecast(ws, plan) -> None:
    row = _title(ws, "Daily contact forecast by service line", "p10 / p90 are the 80% prediction interval.")
    frame = plan.forecast.copy()
    frame["forecast"] = frame["forecast"].round(1)
    frame["p10"] = frame["p10"].round(1)
    frame["p90"] = frame["p90"].round(1)
    _write_frame(ws, frame, start_row=row, number_formats={
        "forecast": "#,##0.0", "p10": "#,##0.0", "p90": "#,##0.0"})


def _sheet_intervals(ws, plan) -> None:
    row = _title(
        ws, "Half-hourly requirement",
        "'Binding constraint' shows whether service level or the occupancy cap drove the number.",
    )
    _write_frame(ws, plan.intervals, start_row=row, number_formats={
        "contacts": "#,##0.0", "on_phone_required": "#,##0.00",
        "rostered_required": "#,##0.00", "service_level": "0.0%", "occupancy": "0.0%"})


def _sheet_scenarios(ws, scenarios: pd.DataFrame) -> None:
    row = _title(
        ws, "Scenario comparison",
        "Every scenario re-runs the full interval sizing. Erlang is non-linear, so FTE does not scale with volume.",
    )
    _write_frame(ws, scenarios, start_row=row, number_formats={
        "peak_required_fte": "#,##0.0", "mean_required_fte": "#,##0.0",
        "worst_gap_fte": "#,##0.0", "total_cost": '"GBP "#,##0',
        "premium_cost": '"GBP "#,##0', "cost_vs_base": '"GBP "#,##0',
        "cost_vs_base_pct": "0.00", "fte_vs_base": "#,##0.0"})


def _sheet_accuracy(ws, plan) -> None:
    row = _title(
        ws, "Forecast accuracy - rolling origin backtest",
        "WAPE is the headline. Bias is signed: a model can have good WAPE and still run high every week.",
    )
    if not plan.selection:
        ws.cell(row=row, column=1, value="Backtest not run for this pack.")
        return
    rows = []
    for key, selected in plan.selection.items():
        for acc in selected.accuracy:
            record = acc.as_row()
            record["chosen"] = "YES" if acc.model == selected.chosen else ""
            rows.append(record)
    frame = pd.DataFrame(rows).sort_values(["line_key", "horizon", "wape_pct"])
    _write_frame(ws, frame, start_row=row, number_formats={
        "wape_pct": "0.00", "mape_pct": "0.00", "bias_pct": "0.00",
        "bias_ratio": "0.000", "actual_total": "#,##0"})


def _sheet_risks(ws, risks: list) -> None:
    row = _title(
        ws, "Early-warning risk register",
        "'Recoverable' means the gap is still outside the recruitment lead time. If it is not, hiring cannot fix it.",
    )
    frame = pd.DataFrame([r.as_row() for r in risks])
    last = _write_frame(ws, frame, start_row=row, number_formats={
        "impact_fte": "#,##0.0", "impact_cost": '"GBP "#,##0', "estimated_saving": '"GBP "#,##0'})
    for r in range(row + 1, last + 1):
        severity = ws.cell(row=r, column=2).value
        colour = {"critical": BAD, "high": WARN, "medium": LIGHT}.get(str(severity), None)
        if colour:
            ws.cell(row=r, column=2).fill = PatternFill("solid", fgColor=colour)
    for col in ("H", "I"):
        ws.column_dimensions[col].width = 70
        for r in range(row + 1, last + 1):
            ws.cell(row=r, column=8 if col == "H" else 9).alignment = Alignment(wrap_text=True, vertical="top")
